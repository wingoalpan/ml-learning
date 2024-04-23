
import sys, os
import argparse
import time
import json as js
import torch
import torch.utils.data as Data
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

import transformer
import mt_lstm
from dataset_P3n9W31 import TP3n9W31Data
from dataset_simple import SimpleData
import bleu

sys.path.append('..\\..\\wingoal_utils')
import common as CM
from common import (
    set_log_file,
    log,
    logs
)
sys.path.append('..\\utils')
import dl_utils

set_log_file(os.path.split(__file__)[-1], timestamp=True)

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

benchmark_test_sentences = [('我 喜欢 黑 猫 。',  'I like black cat .'),
                        ('我 有 一只 小 花猫 。',  'I have a little tabby cat . '),
                        ('我 不 喜欢 白 猫', "I don't like white cat"),
                        ('我 有 两只 黑 猫', 'I have two black cat'),
                        ('我 没有 白 猫', 'I have no white cat'),
                        ('我 喜欢 小 花猫', 'I like little tabby cat'),
                        ('我 喜欢 小 黑 猫', 'I like little black cat'),
                        ('一只 黑 猫', 'a black cat'),
                        ('两只 白 猫', 'two white cat'),
                        ('没有 两只 白 猫', 'have no two white cat'),
                        ('不 喜欢 小 花猫', "don't like little tabby cat"),
                        ('喜欢 小 白 猫', "like little white cat"),
                        ]

benchmark_checkpoints_simple = [50, 100, 200, 300, 400, 600, 800, 1000]
benchmark_epochs_simple = 1000
benchmark_cases_simple = {
    "simple_dropout0.0": {
        'alias': 'simp_drop0',
        'corpora': 'SimpleData',
        'test_sentences': benchmark_test_sentences,
        'num_epochs': 1000,
        'batch_size': 6,
        'dropout': 0.0,
    },
    "simple_dropout0.1": {
        'alias': 'simp_drop01',
        'corpora': 'SimpleData',
        'test_sentences': benchmark_test_sentences,
        'num_epochs': 1000,
        'batch_size': 6,
        'dropout': 0.1,
    },
    "simple_dropout0.3": {
        'alias': 'simp_drop03',
        'corpora': 'SimpleData',
        'test_sentences': benchmark_test_sentences,
        'num_epochs': 1000,
        'batch_size': 6,
        'dropout': 0.3,
    },
    "simple_dropout0.4": {
        'alias': 'simp_drop04',
        'corpora': 'SimpleData',
        'test_sentences': benchmark_test_sentences,
        'num_epochs': 1000,
        'batch_size': 6,
        'dropout': 0.4,
    },
}

benchmark_checkpoints_normal = [50, 100, 150, 200]
benchmark_epochs_normal = 200
benchmark_cases_normal = {
    "norm_dropout0.0": {
        'alias': 'norm_drop0',
        'corpora': 'TP3n9W31Data',
        'num_epochs': 200,
        'batch_size': 48,
        'dropout': 0.0,
    },
    "norm_dropout0.1": {
        'alias': 'norm_drop01',
        'corpora': 'TP3n9W31Data',
        'num_epochs': 200,
        'batch_size': 48,
        'dropout': 0.1,
    },
    "norm_dropout0.3": {
        'alias': 'norm_drop03',
        'corpora': 'TP3n9W31Data',
        'num_epochs': 200,
        'batch_size': 48,
        'dropout': 0.3,
    },
    "norm_dropout0.4": {
        'alias': 'norm_drop04',
        'corpora': 'TP3n9W31Data',
        'num_epochs': 200,
        'batch_size': 48,
        'dropout': 0.4,
    },
}


# check if the trained models for all required epochs existed
def _model_trained(model_name, model_type_name, epochs):
    for epoch in epochs:
        last_states = dl_utils.get_last_state(model_name, model_type_name, max_epoch=epoch)
        if not last_states or not last_states.get('last_epoch', 0) == epoch:
            return False
    return True


def train(cases, model_type, force_retrain=True):
    log('training (model_type: %s) ...' % model_type.MODEL_TYPE)
    train_func = model_type.train

    for name, config in cases.items():
        # if model for all epochs in checkpoints are trained, then no need train again
        if not force_retrain and _model_trained(name, model_type.MODEL_TYPE, benchmark_checkpoints):
            continue

        metrics = cases[name].get('benchmark', {})
        model_metrics = metrics.get(model_type.MODEL_TYPE, {})
        train_metrics = model_metrics.get('train_metric', {})
        # prepare all training required data, model and parameters
        corpora = SimpleData() if config.get('corpora', 'SimpleData') == 'SimpleData' else TP3n9W31Data()
        batch_size = config.get('batch_size', 32)
        dropout = config.get('dropout', 0.0)
        checkpoints = config.get('checkpoints', benchmark_checkpoints)
        num_epochs = config.get('num_epochs', benchmark_epochs)
        loader = Data.DataLoader(corpora, batch_size, True)
        model = model_type.create_model(corpora, name=name, dropout=dropout).to(device)
        # Begin training ...
        start_time = time.time()
        train_func(model, loader, num_epochs,
                   force_retrain=True,
                   checkpoint_interval=0,
                   checkpoints=checkpoints)
        train_seconds = time.time() - start_time
        log('training time consumed: %s secs (model_type: %s, name: %s, epochs: %s) !' %
            (train_seconds, model_type.MODEL_TYPE, name, num_epochs))
        train_metrics['train_seconds'] = train_seconds

        # update the training metrics (training time consumed) to cases
        model_metrics['train_metric'] = train_metrics
        metrics[model_type.MODEL_TYPE] = model_metrics
        cases[name]['benchmark'] = metrics


def translate(cases, model_type):
    log('translating (model_type: %s) ...' % model_type.MODEL_TYPE)
    for name, config in cases.items():
        metrics = cases[name].get('benchmark', {})
        model_metrics = metrics.get(model_type.MODEL_TYPE, {})
        trans_metrics = model_metrics.get('trans_metric', {})
        corpora = SimpleData() if config.get('corpora', 'SimpleData') == 'SimpleData' else TP3n9W31Data()
        latest_epochs = config.get('checkpoints', benchmark_checkpoints)   # [50, 100, 200, 400, 750, 1000]
        model = model_type.create_model(corpora, name=name).to(device)
        # Prepare translation inputs data
        # test_sentences = config.get('test_sentences', benchmark_test_sentences)
        # src_sentences = [item[0] for item in test_sentences]
        # refs_sentences = [[item[1]] for item in test_sentences]
        test_corpora = SimpleData('test') if config.get('corpora', 'SimpleData') == 'SimpleData' else TP3n9W31Data('test')
        src_sentences = test_corpora.sources
        refs_sentences = [[sentence] for sentence in test_corpora.sources]

        enc_inputs = corpora.preprocess(src_sentences)
        # Translating with the models trained by different depth (epoch)
        for latest_epoch in latest_epochs:
            # load the trained model
            last_states = dl_utils.get_last_state(name, model_type.MODEL_TYPE, max_epoch=latest_epoch)
            if not last_states:
                continue
            model.load_state_dict(torch.load(last_states['file_name'], map_location=device))
            # translating the previous inputs ...
            start_time = time.time()
            dec_outputs = model.translate(enc_inputs)
            metric_translate_seconds = time.time() - start_time
            translated_seqs = []

            candidates = []
            for src, dec_output, refs in zip(src_sentences, dec_outputs, refs_sentences):
                prediction = corpora.to_tgt_sentence(dec_output, first=True)
                translated_seqs.append({
                    'source': src,   # 带翻译的原文
                    'translated': prediction,  # 翻译结果
                    'references': refs if len(refs) > 1 else refs[0]   # 翻译标准答案，可以有多个
                })
                candidates.append(prediction)
            # calculate the translation accuracy by BLEU method
            metric_bleu = bleu.bleu(candidates, refs_sentences, 2)
            trans_file_name = '-'.join('trans', model_type.MODEL_TYPE, name, str(latest_epoch)) + '.json'
            CM.save_json(translated_seqs, os.path.join('logs', trans_file_name))
            trans_metrics[str(latest_epoch)] = {'translate_seconds': metric_translate_seconds,
                                                'translate_bleu': metric_bleu,
                                                'translate_seqs': os.path.join('logs', trans_file_name)}
        # update the translation metrics (translating time consumed, bleu etc.) to cases
        model_metrics['trans_metric'] = trans_metrics
        metrics[model_type.MODEL_TYPE] = model_metrics
        cases[name]['benchmark'] = metrics


def report(cases):
    data_train = {}
    data_secs = {'epoch': benchmark_checkpoints}
    data_bleu = {'epoch': benchmark_checkpoints}
    secs_heads = []
    bleu_heads = []

    train_ids = []
    train_secs = []
    for name, config in cases.items():
        for model_type, metrics in config['benchmark'].items():
            case_name_prefix = '-'.join([model_type, name])
            # 提取训练相关度量信息
            train_ids.append(case_name_prefix)
            train_secs.append(metrics.get('train_metric', {}).get('train_seconds', 0.0))
            # 提取翻译相关度量信息
            trans_secs = []
            trans_bleu = []
            for epoch in benchmark_checkpoints:
                trans_secs.append(metrics['trans_metric'].get(str(epoch)).get('translate_seconds'))
                trans_bleu.append(metrics['trans_metric'].get(str(epoch)).get('translate_bleu'))
            data_secs[case_name_prefix + ':' + 'trans_secs'] = trans_secs
            data_bleu[case_name_prefix + ':' + 'trans_bleu'] = trans_bleu
            secs_heads.append(case_name_prefix + ':' + 'trans_secs')
            bleu_heads.append(case_name_prefix + ':' + 'trans_bleu')

    df_secs = pd.DataFrame(data_secs)
    df_secs = df_secs[['epoch'] + secs_heads]
    df_bleu = pd.DataFrame(data_bleu)
    df_bleu = df_bleu[['epoch'] + bleu_heads]

    df_train = pd.DataFrame({'scenario': train_ids, 'train_secs': train_secs})
    df_train = df_train[['scenario', 'train_secs']]

    sheet_train = 'train metric'
    sheet_secs = 'translate secs'
    sheet_bleu = 'translate bleu'
    with pd.ExcelWriter(report_file_name, engine='openpyxl') as writer:
        # train time consumed
        df_train.to_excel(writer, sheet_name=sheet_train, index=False)
        worksheet = writer.sheets[sheet_train]
        worksheet.row_dimensions[1].height = 40
        worksheet.column_dimensions['A'].width = 40
        worksheet.column_dimensions['B'].width = 20
        # translate time consumed
        df_secs.to_excel(writer, sheet_name=sheet_secs, index=False)
        worksheet = writer.sheets[sheet_secs]
        worksheet.row_dimensions[1].height = 40
        for i in range(2, 2 + len(secs_heads)):
            worksheet.column_dimensions[get_column_letter(i)].width = 20
            worksheet.cell(row=1, column=i).alignment = Alignment(wrap_text=True)
        # translation quality (measured by bleu)
        df_bleu.to_excel(writer, sheet_name=sheet_bleu, index=False)
        worksheet = writer.sheets[sheet_bleu]
        worksheet.row_dimensions[1].height = 40
        for i in range(2, 2 + len(bleu_heads)):
            worksheet.column_dimensions[get_column_letter(i)].width = 20
            worksheet.cell(row=1, column=i).alignment = Alignment(wrap_text=True)


benchmark_checkpoints = benchmark_checkpoints_normal
benchmark_epochs = benchmark_epochs_normal
benchmark_cases = benchmark_cases_normal
case_file_name = 'benchmark/benchmark_cases_normal.json'
report_file_name = 'benchmark/report_normal.xlsx'
def main():
    test_cases = CM.load_json(case_file_name)
    test_cases = benchmark_cases if test_cases is None else test_cases
    train(test_cases, transformer, force_retrain=False)
    train(test_cases, mt_lstm, force_retrain=False)
    log('saving train metrics ...')
    CM.save_json(test_cases, case_file_name)
    translate(test_cases, mt_lstm)
    translate(test_cases, transformer)
    log('printing benchmark ...')
    print(js.dumps(test_cases, indent=2, ensure_ascii=False))
    log('saving benchmark ...')
    CM.save_json(test_cases, case_file_name)
    report(test_cases)


def gen_report():
    cases = CM.load_json(case_file_name)
    report(cases)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("params", nargs="*")
    args = parser.parse_args()
    if len(args.params) == 0:
        log('executing function [main] ...')
        main()
    else:
        func = args.params[0]
        if func != 'main':
            set_log_file(os.path.split(__file__)[-1], suffix=func, timestamp=True)
        param_list = args.params[1:]
        log('executing function [%s] ...' % func)
        eval(func)(*param_list)
    log('finish executing function!')

